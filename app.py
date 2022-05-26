from flask import Flask, redirect, render_template, request, Response, redirect
from flask_sqlalchemy import SQLAlchemy
import cv2
from datetime import datetime as dt
import face_recognition
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///register.db'
db = SQLAlchemy(app)
year = ['2022', '2023']
month = [['January', 31], ['February', 28], ['March', 31], ['April', 30], ['May', 31],
         ['June', 30], ['July', 31], ['August', 31], ['September', 30], ['October', 31],
          ['November', 30], ['December', 30]]

# Creation of the table of the attendance of date and month only to be executed once to create the table
# It can also be done manually we just create number of workbooks as 1 per year with each sheet storing a month
# for example here we have create for the 2022 and 2023 workbook 
# for x in range(2):
#     wb = Workbook('Attendance'+year[x]+'.xlsx')
#     for i in range(12):
#         ws = wb.create_sheet(month[i][0])
#         ws.append(['EmailID']+ [str(y+1) for y in range(month[i][1])])
#     wb.save('Attendance'+year[x]+'.xlsx')


class User(db.Model):
    email = db.Column(db.String(200), primary_key=True)
    password = db.Column(db.String(100), nullable=False)
    user_name = db.Column(db.String(100), nullable=False)
    encoding = db.Column(db.String(), nullable=False)
    registration_date = db.Column(db.DateTime, default=dt.utcnow)

    def __repr__(self) -> str:
        return "{}-{}".format(self.email,self.password)


def global_reset():
        global face_encoding
        face_encoding = []
        global user_email
        user_email = ""
        global name_user
        name_user = ""
        global user_password
        user_password = ""
        global match
        match = ""
        global user
        user = None


def add_registration(user_email):
    for x in range(len(year)):
        wb = load_workbook('Attendance'+year[x]+'.xlsx')
        for i in range(12):
            ws = wb[month[i][0]]
            ws.append([user_email])
        wb.save('Attendance'+year[x]+'.xlsx')
    return

# Search for the cell address where to put the attendance
def get_cell_address(user_email):
    list_dt = str(dt.now()).split()
    file_year, month_choice, date = [int(X) for X in list_dt[0].split('-')]
    time = list_dt[1].split('.')[0]
    wb = load_workbook('Attendance'+str(file_year)+'.xlsx')
    ws = wb[month[month_choice-1][0]]
    col_id = get_column_letter(date+1)
    row= 2
    val = None
    while ws['A'+str(row)].value != None:
        if ws['A'+str(row)].value == user_email:
            val = ws[col_id+str(row)].value
            break
        row += 1
    wb.save('Attendance'+str(file_year)+'.xlsx')
    return (col_id + str(row), val)


def set_attendance(user_email):
    list_dt = str(dt.now()).split()
    file_year, month_choice, date = [int(X) for X in list_dt[0].split('-')]
    wb = load_workbook('Attendance'+str(file_year)+'.xlsx')
    ws = wb[month[month_choice-1][0]]
    ws[get_cell_address(user_email)[0]].value = "P"
    wb.save('Attendance'+str(file_year)+'.xlsx')
    return 
    

def  generate_frame ():
    global face_encoding
    global camera
    camera = cv2.VideoCapture(0)
    while True:
        success, frame = camera.read()
        if not success:
            break
        else:
            # frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            frame2 = cv2.resize(frame, (0,0), None, 0.25, 0.25)
            faceLoc = face_recognition.face_locations(frame2)
            face_encoding = face_recognition.face_encodings(frame2)
            if len(faceLoc)!=0:
                faceLoc = faceLoc[0]
                faceLoc2 = [int(x*4) for x in faceLoc]
                frame = cv2.rectangle(frame, (faceLoc2[3], faceLoc2[0]), (faceLoc2[1], faceLoc2[2]), (255, 0, 0), 1)
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
        yield(b'--frame\r\n'b'Content-Type : image/jpeg\r\n\r\n'+frame+b'\r\n')


def  generate_frame_compare():
    count = 0
    global camera
    camera = cv2.VideoCapture(0)
    registered_encoding = [float(x) for x in user.encoding.split()]
    while True:
        success, frame = camera.read()
        if not success:
            break
        else:
            # frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            frame2 = cv2.resize(frame, (0,0), None, 0.25, 0.25)
            faceLoc = face_recognition.face_locations(frame2)
            face_encoding = face_recognition.face_encodings(frame2,faceLoc)
            if len(faceLoc)!=0:
                faceLoc = faceLoc[0]
                faceLoc2 = [int(x*4) for x in faceLoc]
                result = face_recognition.compare_faces([registered_encoding],face_encoding[0],tolerance=0.4)
                if result[0] == True:
                    count +=1
                    frame = cv2.putText(frame,user.user_name, (faceLoc2[3]+6, faceLoc2[2]-6),  cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 1 )
                    if count >= 8:
                        set_attendance(user.email)
                        camera.release()
                        return render_template('completed.html')
                else:
                    frame = cv2.putText(frame,"User Do Not Match", (faceLoc2[3]+6, faceLoc2[2]-6),  cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 1 )
                frame = cv2.rectangle(frame, (faceLoc2[3], faceLoc2[0]), (faceLoc2[1], faceLoc2[2]), (255, 0, 0), 1)
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
        yield(b'--frame\r\n'b'Content-Type : image/jpeg\r\n\r\n'+frame+b'\r\n')


@app.route('/')
def index():
    # The main home poge of the web app is called
    try:
        camera.release()
        if name_user == "" and match==1:
            global_reset()
            # Opens the main page with a text message when the registration is complete
            return render_template('index.html', tasks= 0)
        else:
            global_reset()
            return render_template('index.html',tasks=1)
    except: 
        pass
    global_reset()
    return render_template('index.html', tasks=1)

# Get the registration details of the user 
@app.route('/register', methods=['POST', 'GET'])
def register():
    error = 0
    if request.method == 'POST':
        global user_email
        user_email = request.form['email_id']
        global name_user
        name_user = request.form['user_id']
        global user_password
        user_password = request.form['password']
        global match
        match = request.form['password_match']

        if len(user_email.strip())==0 or len(name_user.strip())==0 or len(user_password.strip())==0 or len(match.strip()) == 0:
            error = 1
            # Error because the fields were empty so ask the user to enter the information again and put a message
            return render_template('register.html', tasks=error)
        elif user_password != match:
            error = 2
             # Error because the password did not match so ask the user to enter the information again and put a message
            return render_template('register.html', tasks=error)
        else:
            try:
                 # Error because the email id already exist so ask the user to enter the information again and put a message
                User.query.get_or_404(user_email)
                error = 3
                return render_template('register.html', tasks=error)
            except:
                return redirect('/registercam')
    else:
        return render_template ('register.html', tasks=error)


@app.route('/registercam', methods=['POST', 'GET'])
def registercam():
    if request.method =="POST":
        if len(face_encoding)==0:
            # if the face is not recognized to come out with a message
            return render_template('registerCam.html', tasks=0)
        else:
            # converted the user face encoding into string to store
            encoding_list = face_encoding[0].tolist()
            encoded_string = " ".join([str(X) for X in encoding_list])
            # create an object to store in the database
            user_register = User(email=user_email, password=user_password, user_name=name_user, encoding=encoded_string)
            # Add the registered data of the user in the database
            db.session.add(user_register)
            db.session.commit()
            add_registration(user_email)
            global_reset()
            # Here match is used as a flag to show that the registration is being done
            global match
            match=1
            return redirect('/')
        
    else:
        return render_template('registerCam.html',tasks=1)


@app.route('/video')
def video():
    # get the camera input in the browser 
    return Response(generate_frame(), mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/login', methods = ['POST', 'GET'])
def login():
    if request.method == 'POST':
        login_id = request.form['login_id']
        login_password = request.form['login_password']
        try:
            global user
            user = User.query.get_or_404(login_id)
            if user.password == login_password:
                return redirect('/login/user')
            else:
                return render_template('login.html', tasks=0)
        except:
            return render_template('login.html', tasks=1)
        
    else:
        return render_template('login.html', tasks=2)


@app.route('/login/user')
def login_user():
    return render_template('user.html', tasks = [user.user_name, user.email])


@app.route('/completed')
def completed():
    return render_template('completed.html')


@app.route('/attendance')
def attendance():
    if user is None: 
        return redirect('/login/user')
    elif get_cell_address(user.email)[1]=='P':
        return redirect('/completed')
    return render_template('put_attendance.html')


@app.route('/video_attend')
def video_attend():
    if user is None:
        return redirect('/login/user')
    # get the camera input in the browser 
    return Response(generate_frame_compare(), mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/view')
def view():
    return render_template('month.html', tasks = [user.user_name, user.email, str(user.registration_date).split()[0]])


@app.route('/month_view/<int:id>')
def month_view(id):
    list_dt = str(dt.now()).split()
    file_year, month_choice, date = [int(X) for X in list_dt[0].split('-')]
    wb = load_workbook('Attendance'+str(file_year)+'.xlsx')
    ws = wb[month[id-1][0]]
    row= 2
    val = []
    while ws['A'+str(row)].value != None:
        print(row)
        if ws['A'+str(row)].value == user.email:
            print("found_row", row)
            break
        row+=1
    for col in range(2, int(month[id-1][1])+2):
        if ws[get_column_letter(col)+str(row)].value =='P':
            val.append("P")
        else:
            val.append("A")
    tasks = [user.user_name, user.email, str(user.registration_date).split()[0], month[id-1], val]
    return render_template('view_attendance.html', tasks=tasks)


if __name__ == '__main__':
    app.run(debug=True)